# =============================================================================
# reports_app/app.py  –  Thinkster Math AI Reports Application
# =============================================================================
# Run with:  python app.py   (inside reports_app/)
# =============================================================================

import os
import io
import re
from datetime import datetime

from dotenv import load_dotenv

# Load .env from the same directory as this file
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

from flask import Flask, render_template, jsonify, request, send_file
from pymongo import MongoClient

# ---------------------------------------------------------------------------
# Config from .env
# ---------------------------------------------------------------------------
MONGO_URI   = os.getenv("MONGO_URI", "")
MONGO_DB    = os.getenv("MONGO_DB", "Thinkster_testing")
SECRET_KEY  = os.getenv("SECRET_KEY", "dev_secret")

if not MONGO_URI:
    raise RuntimeError("MONGO_URI is not set. Check your .env file.")

# ---------------------------------------------------------------------------
# Flask setup
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = SECRET_KEY

# ---------------------------------------------------------------------------
# MongoDB setup
# ---------------------------------------------------------------------------
mongo_client    = MongoClient(MONGO_URI)
db              = mongo_client[MONGO_DB]
report_col      = db["Worksheet_Report"]      # AI analysis results
ws_answers_col  = db["WS_answers"]            # topic names + answer keys
answering_col   = db["Answering_Report"]      # answering run reports

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _topic_map() -> dict:
    """Return {worksheet_id: topic_name} from WS_answers collection."""
    mapping = {}
    for doc in ws_answers_col.find({}, {"worksheetID": 1, "topicName": 1}):
        ws_id  = doc.get("worksheetID", "")
        topic  = doc.get("topicName", "Unknown Topic") or "Unknown Topic"
        if ws_id:
            mapping[ws_id] = topic
    return mapping


def _get_all_reports():
    """
    Aggregate Worksheet_Report by worksheet_id.
    Returns a list of dicts:
      { worksheet_id, topic_name, total_questions,
        issue_count, passed_count, has_issue, questions: [...] }
    """
    topic_map = _topic_map()
    pipeline = [
        {"$group": {
            "_id": "$worksheet_id",
            "total_questions": {"$sum": 1},
            "issue_count":     {"$sum": {"$cond": [{"$eq": ["$status", "Issue"]}, 1, 0]}},
            "passed_count":    {"$sum": {"$cond": [{"$eq": ["$status", "Passed"]}, 1, 0]}},
            "latest_time":     {"$max": "$analysis_time"},
        }},
        {"$sort": {"_id": 1}}
    ]
    rows = list(report_col.aggregate(pipeline))
    result = []
    for row in rows:
        ws_id = row["_id"] or ""
        result.append({
            "worksheet_id":   ws_id,
            "topic_name":     topic_map.get(ws_id, "Unknown Topic"),
            "total_questions":row["total_questions"],
            "issue_count":    row["issue_count"],
            "passed_count":   row["passed_count"],
            "has_issue":      row["issue_count"] > 0,
            "latest_time":    row.get("latest_time", ""),
        })
    return result


def _get_worksheet_detail(ws_id: str):
    """Return full question-level data for one worksheet."""
    docs = list(report_col.find({"worksheet_id": ws_id}).sort("question_number", 1))
    topic_map = _topic_map()
    questions = []
    for d in docs:
        questions.append({
            "question_number": d.get("question_number", 0),
            "image_name":      d.get("image_name", ""),
            "ai_response":     d.get("ai_response", ""),
            "status":          d.get("status", "Unknown"),
            "analysis_time":   d.get("analysis_time", ""),
        })
    return {
        "worksheet_id": ws_id,
        "topic_name":   topic_map.get(ws_id, "Unknown Topic"),
        "questions":    questions,
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/reports")
def api_reports():
    """Return all aggregated report rows."""
    reports = _get_all_reports()
    return jsonify(reports)


@app.route("/api/topics")
def api_topics():
    """Return sorted unique topic names."""
    reports = _get_all_reports()
    topics = sorted({r["topic_name"] for r in reports if r["topic_name"] != "Unknown Topic"})
    return jsonify(topics)


@app.route("/api/topic-summary")
def api_topic_summary():
    """Return summary stats for a given topic."""
    topic = request.args.get("topic", "").strip()
    reports = _get_all_reports()
    if topic:
        rows = [r for r in reports if r["topic_name"] == topic]
    else:
        rows = reports
    total        = len(rows)
    with_issues  = sum(1 for r in rows if r["has_issue"])
    without_issues = total - with_issues
    return jsonify({
        "total":           total,
        "with_issues":     with_issues,
        "without_issues":  without_issues,
    })


@app.route("/api/worksheet/<ws_id>")
def api_worksheet_detail(ws_id):
    """Return detailed question-level data for one worksheet."""
    data = _get_worksheet_detail(ws_id)
    return jsonify(data)


@app.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    """
    Export selected worksheet reports to a PDF-like HTML file for browser printing.
    Accepts JSON body: { "worksheet_ids": [...], "topic": "..." }
    """
    body        = request.get_json(force=True) or {}
    ws_ids      = body.get("worksheet_ids", [])
    topic_filter= body.get("topic", "")

    reports = _get_all_reports()
    if ws_ids:
        rows = [r for r in reports if r["worksheet_id"] in ws_ids]
    elif topic_filter:
        rows = [r for r in reports if r["topic_name"] == topic_filter]
    else:
        rows = reports

    # Build a simple HTML that the browser can print to PDF
    html_parts = [
        "<!DOCTYPE html><html><head>",
        "<meta charset='UTF-8'>",
        "<title>Thinkster Reports Export</title>",
        "<style>",
        "body{font-family:Inter,sans-serif;margin:40px;color:#111}",
        "h1{color:#4F46E5;margin-bottom:8px}",
        "table{width:100%;border-collapse:collapse;margin-top:16px}",
        "th{background:#4F46E5;color:#fff;padding:10px 12px;text-align:left}",
        "td{padding:8px 12px;border-bottom:1px solid #e5e7eb}",
        "tr:nth-child(even) td{background:#f9fafb}",
        ".issue{color:#ef4444;font-weight:600}",
        ".passed{color:#10b981;font-weight:600}",
        "@media print{button{display:none}}",
        "</style></head><body>",
        "<button onclick='window.print()' style='padding:10px 20px;background:#4F46E5;color:#fff;border:none;border-radius:8px;cursor:pointer;margin-bottom:20px'>🖨 Print / Save as PDF</button>",
        f"<h1>Thinkster Math – AI Analysis Reports</h1>",
        f"<p style='color:#6b7280;margin-bottom:4px'>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
        f"<p style='color:#6b7280'>Worksheets: {len(rows)}</p>",
        "<table><thead><tr>",
        "<th>#</th><th>Worksheet ID</th><th>Topic</th>",
        "<th>Questions</th><th>Issues</th><th>Passed</th><th>Status</th><th>Last Analysed</th>",
        "</tr></thead><tbody>",
    ]
    for i, r in enumerate(rows, 1):
        status_cls = "issue" if r["has_issue"] else "passed"
        status_txt = "⚠ Has Issues" if r["has_issue"] else "✔ No Issues"
        html_parts.append(
            f"<tr>"
            f"<td>{i}</td>"
            f"<td>{r['worksheet_id']}</td>"
            f"<td>{r['topic_name']}</td>"
            f"<td>{r['total_questions']}</td>"
            f"<td>{r['issue_count']}</td>"
            f"<td>{r['passed_count']}</td>"
            f"<td class='{status_cls}'>{status_txt}</td>"
            f"<td>{r['latest_time']}</td>"
            f"</tr>"
        )
    html_parts.append("</tbody></table></body></html>")

    html_content = "".join(html_parts)
    buf = io.BytesIO(html_content.encode("utf-8"))
    buf.seek(0)
    filename = f"thinkster_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    return send_file(buf, mimetype="text/html",
                     as_attachment=True, download_name=filename)


@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    """Export selected worksheet reports to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    body        = request.get_json(force=True) or {}
    ws_ids      = body.get("worksheet_ids", [])
    topic_filter= body.get("topic", "")

    reports = _get_all_reports()
    if ws_ids:
        rows = [r for r in reports if r["worksheet_id"] in ws_ids]
    elif topic_filter:
        rows = [r for r in reports if r["topic_name"] == topic_filter]
    else:
        rows = reports

    wb = Workbook()
    ws = wb.active
    ws.title = "AI Analysis Reports"

    # Header styles
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    headers = ["#", "Worksheet ID", "Topic", "Questions", "Issues", "Passed", "Status", "Last Analysed"]
    col_widths = [5, 30, 40, 12, 10, 10, 16, 24]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28

    issue_font  = Font(bold=True, color="EF4444")
    passed_font = Font(bold=True, color="10B981")

    for i, r in enumerate(rows, 1):
        row_num = i + 1
        status  = "⚠ Has Issues" if r["has_issue"] else "✔ No Issues"
        values  = [i, r["worksheet_id"], r["topic_name"],
                   r["total_questions"], r["issue_count"], r["passed_count"],
                   status, r["latest_time"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center" if col_idx in (1,4,5,6) else "left",
                                       vertical="center", wrap_text=True)
            cell.border = thin_border
            if col_idx == 7:
                cell.font = issue_font if r["has_issue"] else passed_font
        ws.row_dimensions[row_num].height = 20

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16
    summary_data = [
        ("Total Worksheets",        len(rows)),
        ("With Issues",             sum(1 for r in rows if r["has_issue"])),
        ("Without Issues",          sum(1 for r in rows if not r["has_issue"])),
        ("Total Questions Analysed",sum(r["total_questions"] for r in rows)),
        ("Total Issues Found",      sum(r["issue_count"] for r in rows)),
        ("Total Passed",            sum(r["passed_count"] for r in rows)),
        ("Exported At",             datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_idx, (label, val) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"thinkster_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("=" * 60)
    print("  Thinkster Math – Reports Application")
    print("  http://127.0.0.1:5050")
    print("=" * 60)
    app.run(debug=True, host="0.0.0.0", port=5050)
